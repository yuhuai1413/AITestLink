from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.routers.deps import get_current_user, get_test_design_service, get_project_service
from app.services.export_format import current_export_datetime, format_export_datetime
from app.services.test_design_service import TestDesignService
from app.services.project_service import ProjectService
from app.schemas.test_point import TestPointCreate, TestPointUpdate

router = APIRouter()


def _estimate_row_height(values: list[str]) -> int:
    long_text = max((len(value or "") for value in values), default=0)
    line_count = max((str(value or "").count("\n") + 1 for value in values), default=1)
    if long_text > 220 or line_count >= 5:
        return 96
    if long_text > 140 or line_count >= 4:
        return 78
    if long_text > 80 or line_count >= 3:
        return 60
    if long_text > 40 or line_count >= 2:
        return 44
    return 32


def _point_value(item: dict, key: str, index: int) -> str:
    values = {
        "index": str(index + 1),
        "pointCode": item.get("pointCode") or "",
        "module": item.get("module") or "",
        "type": item.get("type") or "",
        "title": item.get("title") or "",
        "description": item.get("description") or "",
        "priority": item.get("priority") or "",
        "automatable": "是" if item.get("automatable") else "否",
        "reviewStatus": item.get("reviewStatus") or "待评审",
        "validityStatus": item.get("validityStatus") or "有效",
        "invalidReason": item.get("invalidReason") or "",
        "createdAt": format_export_datetime(item.get("createdAt")),
        "updatedAt": format_export_datetime(item.get("updatedAt")),
    }
    return str(values.get(key, ""))


def _build_test_points_xlsx(project_name: str, rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    columns = [
        ("index", "序号", 6, "center"),
        ("pointCode", "测试点编号", 18, "center"),
        ("module", "模块", 16, "center"),
        ("type", "类型", 14, "center"),
        ("title", "测试点", 32, "left"),
        ("description", "描述", 44, "left"),
        ("priority", "优先级", 10, "center"),
        ("automatable", "是否自动化", 12, "center"),
        ("reviewStatus", "评审状态", 12, "center"),
        ("validityStatus", "数据状态", 12, "center"),
        ("invalidReason", "失效原因", 28, "left"),
        ("createdAt", "生成时间", 20, "center"),
        ("updatedAt", "更新时间", 20, "center"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "测试点"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    last_col = len(columns)
    last_col_letter = get_column_letter(last_col)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = f"{project_name} - 测试点"
    ws["A1"].font = Font(name="微软雅黑", size=15, bold=True, color="1f2937")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    generated_at = current_export_datetime()
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"] = f"导出时间：{generated_at}    测试点数量：{len(rows)}"
    ws["A2"].font = Font(name="微软雅黑", size=10, color="64748b")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    header_fill = PatternFill("solid", fgColor="E7E6E6")
    border_side = Side(style="thin", color="D9D9D9")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for col_index, (_, label, width, _) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_index, value=label)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="1f2937")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_index)].width = width
    ws.row_dimensions[3].height = 26

    center_keys = {key for key, _, _, align in columns if align == "center"}
    for row_index, item in enumerate(rows, start=4):
        row_values = [_point_value(item, key, row_index - 4) for key, *_ in columns]
        for col_index, (key, _, _, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=row_values[col_index - 1])
            cell.font = Font(name="微软雅黑", size=10, color="111827")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if key in center_keys else "left",
                vertical="center",
                wrap_text=True,
            )
        ws.row_dimensions[row_index].height = _estimate_row_height(row_values)

    ws.auto_filter.ref = f"A3:{last_col_letter}{max(3, len(rows) + 3)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


class GenerateTestPointsRequest(BaseModel):
    requirement_ids: list[str]


@router.get("/projects/{project_id}/test-points")
async def list_test_points(
    project_id: str,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    return await service.list_test_points(project_id)


@router.get("/projects/{project_id}/test-points/export")
async def export_test_points(
    project_id: str,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
    project_service: ProjectService = Depends(get_project_service),
):
    project = await project_service._verify_project_owner(project_id, user["sub"])
    points = await service.list_test_points(project_id)
    content = _build_test_points_xlsx(project.name or "未命名项目", points)
    filename = f"{project.name or '未命名项目'}-测试点.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.post("/projects/{project_id}/test-points", status_code=201)
async def create_test_point(
    project_id: str,
    data: TestPointCreate,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    # 直接创建单个测试点
    from app.models.test_point import TestPoint
    import uuid
    tp = TestPoint(
        id=str(uuid.uuid4()),
        point_code=await service._generate_point_code(project_id, data.module),
        project_id=project_id,
        module=data.module,
        type=data.type,
        title=data.title,
        description=data.description,
        priority=data.priority,
        automatable=data.automatable,
    )
    service.db.add(tp)
    await service.db.commit()
    await service.db.refresh(tp)
    return service._to_dict(tp)


@router.post("/projects/{project_id}/test-points/generate")
async def generate_test_points(
    project_id: str,
    data: GenerateTestPointsRequest,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    return await service.generate_test_points(project_id, data.requirement_ids, user["sub"])


@router.put("/test-points/{tp_id}")
async def update_test_point(
    tp_id: str,
    data: TestPointUpdate,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
):
    tp = await service.update_test_point(tp_id, data)
    if not tp:
        raise HTTPException(status_code=404, detail="Test point not found")
    return tp


@router.delete("/test-points/{tp_id}")
async def delete_test_point(
    tp_id: str,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
):
    raise HTTPException(
        status_code=400,
        detail="测试点属于测试链路中间产物，不允许单独删除；如需调整，请修改评审状态或重新生成测试点。",
    )


class BatchReviewRequest(BaseModel):
    ids: list[str]
    status: str


@router.post("/test-points/batch-review")
async def batch_review_test_points(
    data: BatchReviewRequest,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
):
    count = await service.batch_update_review(data.ids, data.status)
    return {"ok": True, "updated": count}
