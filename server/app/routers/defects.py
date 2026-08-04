import re
from io import BytesIO
from datetime import datetime, timezone
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func, text

from app.routers.deps import get_current_user, get_project_service
from app.database import async_session
from app.models.defect import Defect
from app.schemas.defect import DefectCreate, DefectUpdate
from app.services.project_service import ProjectService

router = APIRouter()


class BatchStatusRequest(BaseModel):
    ids: list[str]
    status: str


def _defect_to_dict(row) -> dict:
    return {
        "id": row.id,
        "projectId": row.project_id,
        "defectCode": row.defect_code,
        "title": row.title,
        "description": row.description or "",
        "severity": row.severity,
        "priority": row.priority,
        "status": row.status,
        "module": row.module,
        "category": row.category,
        "source": row.source or "手工",
        "testCaseId": row.test_case_id,
        "scriptId": row.script_id,
        "executionRunId": row.execution_run_id,
        "stepsToReproduce": row.steps_to_reproduce,
        "expectedResult": row.expected_result,
        "actualResult": row.actual_result,
        "environmentInfo": row.environment_info,
        "reporter": row.reporter,
        "assignee": row.assignee,
        "remark": row.remark,
        "screenshotUrl": row.screenshot_url or "",
        "foundAt": row.found_at.isoformat() if row.found_at else None,
        "resolvedAt": row.resolved_at.isoformat() if row.resolved_at else None,
        "createdAt": row.created_at.isoformat() if row.created_at else None,
        "updatedAt": row.updated_at.isoformat() if row.updated_at else None,
    }


async def _next_defect_code(project_id: str, db) -> str:
    result = await db.execute(
        text("SELECT defect_code FROM defects WHERE project_id = :pid ORDER BY created_at DESC LIMIT 1"),
        {"pid": project_id},
    )
    row = result.fetchone()
    if row and row[0]:
        m = re.search(r"(\d+)$", row[0])
        if m:
            return f"BUG-{int(m.group(1)) + 1:04d}"
    return "BUG-0001"


def _build_defects_xlsx(project_name: str, rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    columns = [
        ("defectCode", "缺陷编号", 16, "center"),
        ("title", "缺陷标题", 36, "left"),
        ("severity", "严重程度", 12, "center"),
        ("priority", "优先级", 10, "center"),
        ("status", "状态", 12, "center"),
        ("module", "模块", 16, "center"),
        ("category", "缺陷类型", 14, "center"),
        ("stepsToReproduce", "复现步骤", 44, "left"),
        ("expectedResult", "期望结果", 30, "left"),
        ("actualResult", "实际结果", 30, "left"),
        ("environmentInfo", "环境信息", 20, "left"),
        ("reporter", "发现人", 12, "center"),
        ("assignee", "指派人", 12, "center"),
        ("remark", "备注", 24, "left"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "缺陷列表"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    last_col = len(columns)
    last_col_letter = get_column_letter(last_col)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = f"{project_name} - 缺陷列表"
    ws["A1"].font = Font(name="微软雅黑", size=15, bold=True, color="1f2937")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    from app.services.export_format import current_export_datetime
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"] = f"导出时间：{current_export_datetime()}    缺陷数量：{len(rows)}"
    ws["A2"].font = Font(name="微软雅黑", size=10, color="64748b")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    header_fill = PatternFill("solid", fgColor="E7E6E6")
    border_side = Side(style="thin", color="D9D9D9")
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    body_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    center_keys = {key for key, _, _, align in columns if align == "center"}

    for col_index, (_, label, width, _) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_index, value=label)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="1f2937")
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_index)].width = width
    ws.row_dimensions[3].height = 26

    for row_index, d in enumerate(rows, start=4):
        for col_index, (key, _, _, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=str(d.get(key, "") or ""))
            cell.font = Font(name="微软雅黑", size=10, color="111827")
            cell.border = body_border
            cell.alignment = Alignment(
                horizontal="center" if key in center_keys else "left",
                vertical="center",
                wrap_text=True,
            )
        ws.row_dimensions[row_index].height = 32

    ws.auto_filter.ref = f"A3:{last_col_letter}{max(3, len(rows) + 3)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


@router.get("/projects/{project_id}/defects")
async def list_defects(
    project_id: str,
    user: dict = Depends(get_current_user),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    async with async_session() as db:
        result = await db.execute(
            select(Defect).where(Defect.project_id == project_id).order_by(Defect.created_at.desc())
        )
        rows = result.scalars().all()
        return [_defect_to_dict(r) for r in rows]


@router.post("/projects/{project_id}/defects", status_code=201)
async def create_defect(
    project_id: str,
    data: DefectCreate,
    user: dict = Depends(get_current_user),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    async with async_session() as db:
        defect_code = await _next_defect_code(project_id, db)
        d = Defect(
            project_id=project_id,
            defect_code=defect_code,
            title=data.title,
            severity=data.severity,
            priority=data.priority,
            status=data.status,
            module=data.module,
            category=data.category,
            source=data.source,
            test_case_id=data.test_case_id,
            script_id=data.script_id,
            execution_run_id=data.execution_run_id,
            steps_to_reproduce=data.steps_to_reproduce,
            expected_result=data.expected_result,
            actual_result=data.actual_result,
            environment_info=data.environment_info,
            reporter=data.reporter,
            assignee=data.assignee,
            remark=data.remark,
            description=data.description,
            screenshot_url=data.screenshot_url,
        )
        db.add(d)
        await db.commit()
        await db.refresh(d)
        return _defect_to_dict(d)


@router.put("/defects/{defect_id}")
async def update_defect(
    defect_id: str,
    data: DefectUpdate,
    user: dict = Depends(get_current_user),
):
    async with async_session() as db:
        result = await db.execute(select(Defect).where(Defect.id == defect_id))
        d = result.scalar_one_or_none()
        if not d:
            raise HTTPException(status_code=404, detail="缺陷不存在")
        update_data = data.model_dump(exclude_unset=True)
        field_map = {
            "test_case_id": "test_case_id",
            "steps_to_reproduce": "steps_to_reproduce",
            "expected_result": "expected_result",
            "actual_result": "actual_result",
            "environment_info": "environment_info",
        }
        for key, value in update_data.items():
            model_key = field_map.get(key, key)
            setattr(d, model_key, value)
        if data.status == "已修复" and not d.resolved_at:
            d.resolved_at = datetime.now(timezone.utc)
        d.updated_at = datetime.now(timezone.utc)
        await db.commit()
        await db.refresh(d)
        return _defect_to_dict(d)


@router.delete("/defects/{defect_id}")
async def delete_defect(
    defect_id: str,
    user: dict = Depends(get_current_user),
):
    async with async_session() as db:
        result = await db.execute(select(Defect).where(Defect.id == defect_id))
        d = result.scalar_one_or_none()
        if not d:
            raise HTTPException(status_code=404, detail="缺陷不存在")
        await db.delete(d)
        await db.commit()
        return {"ok": True}


@router.post("/defects/batch-status")
async def batch_update_status(
    data: BatchStatusRequest,
    user: dict = Depends(get_current_user),
):
    async with async_session() as db:
        result = await db.execute(select(Defect).where(Defect.id.in_(data.ids)))
        rows = result.scalars().all()
        for d in rows:
            d.status = data.status
            if data.status == "已修复" and not d.resolved_at:
                d.resolved_at = datetime.now(timezone.utc)
            d.updated_at = datetime.now(timezone.utc)
        await db.commit()
        return {"ok": True, "updated": len(rows)}


@router.get("/projects/{project_id}/defects/export")
async def export_defects(
    project_id: str,
    user: dict = Depends(get_current_user),
    project_service: ProjectService = Depends(get_project_service),
):
    project = await project_service._verify_project_owner(project_id, user["sub"])
    async with async_session() as db:
        result = await db.execute(
            select(Defect).where(Defect.project_id == project_id).order_by(Defect.created_at.desc())
        )
        rows = result.scalars().all()
        data = [_defect_to_dict(r) for r in rows]

    content = _build_defects_xlsx(project.name or "未命名项目", data)
    filename = f"{project.name or '未命名项目'}-缺陷列表.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.get("/projects/{project_id}/defects/stats")
async def defect_stats(
    project_id: str,
    user: dict = Depends(get_current_user),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    async with async_session() as db:
        result = await db.execute(
            select(Defect).where(Defect.project_id == project_id)
        )
        rows = result.scalars().all()
        total = len(rows)
        by_severity = {}
        by_status = {}
        by_module = {}
        by_category = {}
        by_source = {}
        for r in rows:
            by_severity[r.severity] = by_severity.get(r.severity, 0) + 1
            by_status[r.status] = by_status.get(r.status, 0) + 1
            by_module[r.module] = by_module.get(r.module, 0) + 1
            by_category[r.category] = by_category.get(r.category, 0) + 1
            src = r.source or "手工"
            by_source[src] = by_source.get(src, 0) + 1
        return {
            "total": total,
            "bySeverity": by_severity,
            "byStatus": by_status,
            "byModule": by_module,
            "byCategory": by_category,
            "bySource": by_source,
            "autoCount": by_source.get("自动化", 0),
            "openCount": sum(1 for r in rows if r.status not in ("已关闭", "已验证")),
            "closedCount": sum(1 for r in rows if r.status in ("已关闭", "已验证")),
        }


@router.post("/upload-image")
async def upload_editor_image(
    file: UploadFile,
    user: dict = Depends(get_current_user),
):
    """Upload an image for use in rich text editors."""
    import os, uuid
    from app.config import settings
    
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="只能上传图片文件")
    
    ext = os.path.splitext(file.filename or "image.png")[1] or ".png"
    filename = f"editor_{uuid.uuid4().hex}{ext}"
    
    upload_dir = os.path.join(settings.UPLOAD_DIR, "editor-images")
    os.makedirs(upload_dir, exist_ok=True)
    
    content_bytes = await file.read()
    if len(content_bytes) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=400, detail="图片大小不能超过10MB")
    
    filepath = os.path.join(upload_dir, filename)
    with open(filepath, "wb") as f:
        f.write(content_bytes)
    
    return {"url": f"/uploads/editor-images/{filename}"}
