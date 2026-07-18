from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.routers.deps import get_current_user, get_test_design_service, get_project_service
from app.services.test_design_service import TestDesignService
from app.services.project_service import ProjectService
from app.schemas.test_case import TestCaseCreate, TestCaseUpdate

router = APIRouter()


class GenerateTestCasesRequest(BaseModel):
    test_point_ids: list[str]


class BatchStatusRequest(BaseModel):
    ids: list[str]
    status: str


def _format_steps_for_export(steps: str | None) -> str:
    import re

    text = (steps or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    text = re.sub(r"[ \t]*\n+[ \t]*", "\n", text)
    text = re.sub(r"\s+(?=(?:步骤\s*\d+|第[一二三四五六七八九十百千万]+步)\s*[:：])", "\n", text)
    text = re.sub(r"\s+(?=\d+\s*[.、]\s*)", "\n", text)
    return text.strip()


def _case_value(tc: dict, key: str, index: int) -> str:
    actual = (tc.get("actualResult") or "").strip()
    expected = (tc.get("expectedResult") or "").strip()
    passed = "通过" if actual and expected and actual == expected else ("未通过" if actual else (tc.get("passed") or "未执行"))
    values = {
        "index": str(index + 1),
        "caseCode": tc.get("caseCode") or "",
        "module": tc.get("module") or "",
        "testType": tc.get("testType") or "功能测试",
        "feature": tc.get("feature") or "",
        "title": tc.get("title") or "",
        "priority": tc.get("priority") or "",
        "targetPlatform": tc.get("targetPlatform") or "PC",
        "testUrl": tc.get("testUrl") or "未配置",
        "requiredRole": tc.get("requiredRole") or "无",
        "precondition": tc.get("precondition") or "",
        "steps": _format_steps_for_export(tc.get("steps") or ""),
        "testData": tc.get("testData") or "",
        "expectedResult": tc.get("expectedResult") or "",
        "actualResult": tc.get("actualResult") or "",
        "passed": passed,
        "reviewStatus": tc.get("reviewStatus") or "待评审",
        "automation": "是" if tc.get("automation") == "是" else "否",
        "remark": tc.get("remark") or "",
    }
    return str(values.get(key, ""))


def _estimate_row_height(values: list[str]) -> int:
    long_text = max((len(value or "") for value in values), default=0)
    line_count = max((str(value or "").count("\n") + 1 for value in values), default=1)
    if long_text > 220 or line_count >= 5:
        return 96
    if long_text > 140 or line_count >= 4:
        return 78
    if long_text > 80 or line_count >= 3:
        return 60
    if long_text > 40 or line_count >= 2:
        return 44
    return 32


def _build_test_cases_xlsx(project_name: str, rows: list[dict], export_type: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    columns = [
        ("index", "序号", 6, "center"),
        ("caseCode", "用例编号", 18, "center"),
        ("module", "模块", 16, "center"),
        ("testType", "测试类型", 12, "center"),
        ("feature", "测试点", 26, "left"),
        ("title", "用例标题", 32, "left"),
        ("priority", "优先级", 10, "center"),
        ("targetPlatform", "测试端", 10, "center"),
        ("testUrl", "测试地址", 34, "left"),
        ("requiredRole", "所需角色", 14, "center"),
        ("precondition", "前置条件", 24, "left"),
        ("steps", "测试步骤", 44, "left"),
        ("testData", "测试数据", 24, "left"),
        ("expectedResult", "预期结果", 36, "left"),
        ("actualResult", "实测结果", 30, "left"),
        ("passed", "执行结果", 12, "center"),
        ("reviewStatus", "评审状态", 12, "center"),
        ("automation", "自动化", 10, "center"),
        ("remark", "备注", 24, "left"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    last_col = len(columns)
    last_col_letter = get_column_letter(last_col)
    title = f"{project_name} - {'全部测试用例' if export_type == 'all' else '手动测试用例'}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = title
    ws["A1"].font = Font(name="微软雅黑", size=15, bold=True, color="1f2937")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    from datetime import datetime
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"] = f"导出时间：{generated_at}    用例数量：{len(rows)}"
    ws["A2"].font = Font(name="微软雅黑", size=10, color="64748b")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    header_fill = PatternFill("solid", fgColor="E7E6E6")
    border_side = Side(style="thin", color="D9D9D9")
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    body_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for col_index, (_, label, width, _) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_index, value=label)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="1f2937")
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_index)].width = width
    ws.row_dimensions[3].height = 26

    center_keys = {key for key, _, _, align in columns if align == "center"}
    for row_index, tc in enumerate(rows, start=4):
        row_values = [_case_value(tc, key, row_index - 4) for key, *_ in columns]
        for col_index, (key, _, _, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=row_values[col_index - 1])
            cell.font = Font(name="微软雅黑", size=10, color="111827")
            cell.border = body_border
            cell.alignment = Alignment(
                horizontal="center" if key in center_keys else "left",
                vertical="center",
                wrap_text=True,
            )
        ws.row_dimensions[row_index].height = _estimate_row_height(row_values)

    ws.auto_filter.ref = f"A3:{last_col_letter}{max(3, len(rows) + 3)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


@router.get("/projects/{project_id}/test-cases")
async def list_test_cases(
    project_id: str,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    return await service.list_test_cases(project_id)


@router.get("/projects/{project_id}/test-cases/export")
async def export_test_cases(
    project_id: str,
    type: str = Query("all", pattern="^(all|manual)$"),
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
    project_service: ProjectService = Depends(get_project_service),
):
    project = await project_service._verify_project_owner(project_id, user["sub"])
    cases = await service.list_test_cases(project_id)
    if type == "manual":
        cases = [tc for tc in cases if tc.get("automation") != "是"]

    content = _build_test_cases_xlsx(project.name or "未命名项目", cases, type)
    suffix = "全部测试用例" if type == "all" else "手动测试用例"
    filename = f"{project.name or '未命名项目'}-{suffix}.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.post("/projects/{project_id}/test-cases", status_code=201)
async def create_test_case(
    project_id: str,
    data: TestCaseCreate,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    # 直接创建单个测试用例
    from app.models.test_case import TestCase
    import uuid
    tc = TestCase(
        id=str(uuid.uuid4()),
        project_id=project_id,
        test_point_id=data.test_point_id,
        requirement_id=data.requirement_id,
        case_code=data.case_code,
        module=data.module,
        feature=data.feature,
        title=data.title,
        priority=data.priority,
        precondition=data.precondition,
        steps=data.steps,
        test_data=data.test_data,
        expected_result=data.expected_result,
        environment_id=data.environment_id,
        target_platform=data.target_platform,
        test_url=data.test_url,
        required_role=data.required_role,
        automation=data.automation,
        review_status=data.review_status,
        remark=data.remark,
    )
    service.db.add(tc)
    await service.db.commit()
    await service.db.refresh(tc)
    return service._to_dict(tc)


@router.post("/projects/{project_id}/test-cases/generate")
async def generate_test_cases(
    project_id: str,
    data: GenerateTestCasesRequest,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    return await service.generate_test_cases(project_id, data.test_point_ids, user["sub"])


@router.put("/test-cases/{tc_id}")
async def update_test_case(
    tc_id: str,
    data: TestCaseUpdate,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
):
    tc = await service.update_test_case(tc_id, data)
    if not tc:
        raise HTTPException(status_code=404, detail="Test case not found")
    return tc


@router.delete("/test-cases/{tc_id}")
async def delete_test_case(
    tc_id: str,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
):
    raise HTTPException(
        status_code=400,
        detail="测试用例属于测试链路中间产物，不允许单独删除；如需调整，请修改评审状态或重新生成测试用例。",
    )


@router.post("/test-cases/batch-status")
async def batch_update_status(
    data: BatchStatusRequest,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
):
    count = await service.batch_update_status(data.ids, data.status)
    return {"ok": True, "updated": count}


@router.post("/test-cases/batch-review")
async def batch_review_test_cases(
    data: BatchStatusRequest,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
):
    count = await service.batch_update_review(data.ids, data.status)
    return {"ok": True, "updated": count}


@router.get("/projects/{project_id}/coverage")
async def get_coverage(
    project_id: str,
    user: dict = Depends(get_current_user),
    service: TestDesignService = Depends(get_test_design_service),
):
    return await service.get_coverage(project_id)
