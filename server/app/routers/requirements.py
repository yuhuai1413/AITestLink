from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.routers.deps import get_current_user, get_document_service, get_project_service
from app.services.document_service import DocumentService
from app.services.export_format import current_export_datetime, format_export_datetime
from app.services.project_service import ProjectService
from app.schemas.requirement import RequirementUpdate

router = APIRouter()


def _estimate_row_height(values: list[str]) -> int:
    long_text = max((len(value or "") for value in values), default=0)
    line_count = max((str(value or "").count("\n") + 1 for value in values), default=1)
    if long_text > 220 or line_count >= 5:
        return 96
    if long_text > 140 or line_count >= 4:
        return 78
    if long_text > 80 or line_count >= 3:
        return 60
    if long_text > 40 or line_count >= 2:
        return 44
    return 32


def _requirement_value(item: dict, key: str, index: int) -> str:
    values = {
        "index": str(index + 1),
        "reqId": item.get("reqId") or "",
        "module": item.get("module") or "",
        "feature": item.get("feature") or "",
        "source": item.get("source") or "",
        "risk": item.get("risk") or "",
        "rule": item.get("rule") or "",
        "question": item.get("question") or "",
        "clarificationStatus": item.get("clarificationStatus") or ("已确认" if item.get("confirmed") else "待确认"),
        "clarificationAnswer": item.get("clarificationAnswer") or "",
        "reviewStatus": item.get("reviewStatus") or "待评审",
        "validityStatus": item.get("validityStatus") or "有效",
        "invalidReason": item.get("invalidReason") or "",
        "createdAt": format_export_datetime(item.get("createdAt")),
        "updatedAt": format_export_datetime(item.get("updatedAt")),
    }
    return str(values.get(key, ""))


def _build_requirements_xlsx(project_name: str, rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    columns = [
        ("index", "序号", 6, "center"),
        ("reqId", "需求编号", 18, "center"),
        ("module", "模块", 16, "center"),
        ("feature", "功能点", 24, "left"),
        ("source", "来源", 20, "left"),
        ("risk", "风险等级", 10, "center"),
        ("rule", "业务规则", 42, "left"),
        ("question", "待确认问题", 36, "left"),
        ("clarificationStatus", "确认状态", 12, "center"),
        ("clarificationAnswer", "确认结论", 34, "left"),
        ("reviewStatus", "评审状态", 12, "center"),
        ("validityStatus", "数据状态", 12, "center"),
        ("invalidReason", "失效原因", 28, "left"),
        ("createdAt", "生成时间", 20, "center"),
        ("updatedAt", "更新时间", 20, "center"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "需求列表"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    last_col = len(columns)
    last_col_letter = get_column_letter(last_col)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws["A1"] = f"{project_name} - 需求列表"
    ws["A1"].font = Font(name="微软雅黑", size=15, bold=True, color="1f2937")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    generated_at = current_export_datetime()
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws["A2"] = f"导出时间：{generated_at}    需求数量：{len(rows)}"
    ws["A2"].font = Font(name="微软雅黑", size=10, color="64748b")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    header_fill = PatternFill("solid", fgColor="E7E6E6")
    border_side = Side(style="thin", color="D9D9D9")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    for col_index, (_, label, width, _) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_index, value=label)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="1f2937")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_index)].width = width
    ws.row_dimensions[3].height = 26

    center_keys = {key for key, _, _, align in columns if align == "center"}
    for row_index, item in enumerate(rows, start=4):
        row_values = [_requirement_value(item, key, row_index - 4) for key, *_ in columns]
        for col_index, (key, _, _, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=row_values[col_index - 1])
            cell.font = Font(name="微软雅黑", size=10, color="111827")
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if key in center_keys else "left",
                vertical="center",
                wrap_text=True,
            )
        ws.row_dimensions[row_index].height = _estimate_row_height(row_values)

    ws.auto_filter.ref = f"A3:{last_col_letter}{max(3, len(rows) + 3)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


@router.get("/projects/{project_id}/requirements")
async def list_requirements(
    project_id: str,
    user: dict = Depends(get_current_user),
    service: DocumentService = Depends(get_document_service),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    return await service.get_requirements(project_id)


@router.get("/projects/{project_id}/requirements/export")
async def export_requirements(
    project_id: str,
    user: dict = Depends(get_current_user),
    service: DocumentService = Depends(get_document_service),
    project_service: ProjectService = Depends(get_project_service),
):
    project = await project_service._verify_project_owner(project_id, user["sub"])
    requirements = await service.get_requirements(project_id)
    content = _build_requirements_xlsx(project.name or "未命名项目", requirements)
    filename = f"{project.name or '未命名项目'}-需求列表.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@router.get("/projects/{project_id}/requirements/search")
async def search_requirements(
    project_id: str,
    q: str,
    user: dict = Depends(get_current_user),
    service: DocumentService = Depends(get_document_service),
    project_service: ProjectService = Depends(get_project_service),
):
    await project_service._verify_project_owner(project_id, user["sub"])
    return await service.search_requirements(project_id, q)


@router.put("/requirements/{req_id}")
async def update_requirement(
    req_id: str,
    data: RequirementUpdate,
    user: dict = Depends(get_current_user),
    service: DocumentService = Depends(get_document_service),
):
    req = await service.update_requirement(req_id, data)
    if not req:
        raise HTTPException(status_code=404, detail="Requirement not found")
    return req


@router.delete("/requirements/{req_id}")
async def delete_requirement(
    req_id: str,
    user: dict = Depends(get_current_user),
    service: DocumentService = Depends(get_document_service),
):
    raise HTTPException(
        status_code=400,
        detail="需求属于测试链路中间产物，不允许单独删除；如需调整，请修改评审状态或重新解析输入资料。",
    )


class BatchConfirmRequest(BaseModel):
    ids: list[str]
    confirmed: bool


@router.post("/requirements/batch-confirm")
async def batch_confirm_requirements(
    data: BatchConfirmRequest,
    user: dict = Depends(get_current_user),
    service: DocumentService = Depends(get_document_service),
):
    count = await service.batch_confirm(data.ids, data.confirmed)
    return {"ok": True, "updated": count}
